#!/usr/bin/env python3
# coding: utf-8

import requests
from lxml import html
import openpyxl
import collections
import sys
import dbUtil


def update_veh_type(veh_type_obj):
    sql = "select * from yw_standed_veh_model v where v.product_no = '%s'" % veh_type_obj.product_no
    conn = dbUtil.get_conn()
    cursor = dbUtil.get_cursor(conn)
    res = dbUtil.fetchall(cursor, sql)
    try:
        if len(res) > 0:
            updateSql = """
                    update yw_standed_veh_model set product_name = '%s', 
                                                    announce_batch = 30, 
                                                    oem = '%s',
                                                    veh_type = '%s',
                                                    veh_grade = '%s',
                                                    total_weight = '%s',
                                                    fuel_type = '%s',
                                                    update_time = now(),
                                                    update_by = 'stone'
                    where product_no = '%s'
                """ % (veh_type_obj.product_name, veh_type_obj.oem, veh_type_obj.veh_type, veh_type_obj.veh_grade,
                       veh_type_obj.total_weight, veh_type_obj.fuel_type, veh_type_obj.product_no)

            print("update")
        else:
            updateSql = """
                    insert into yw_standed_veh_model(product_no,
                                                    product_name,
                                                    announce_batch,
                                                    oem,
                                                    veh_type,
                                                    veh_grade,
                                                    total_weight,
                                                    fuel_type,
                                                    create_by, 
                                                    create_time)
                                                    values('%s', '%s', '%s', '%s', '%s', '%s', %s, '%s', 'stone', now())
                """ % (veh_type_obj.product_no, veh_type_obj.product_name, 30, veh_type_obj.oem, veh_type_obj.veh_type,
                       veh_type_obj.veh_grade, veh_type_obj.total_weight, veh_type_obj.fuel_type)
            print("insert")
        dbUtil.update(cursor, updateSql)
        conn.commit()
        print("commit success")
    except:
        print(sys.exc_info())
        print("commit error")
        conn.rollback()
    conn.close()


if __name__ == '__main__':
    # fetchall(get_cursor(), "select version()")
    product_no = "YBL61253H12QCE1"
    product_name = "客车"
    oem = "扬州亚星客车股份有限公司"
    veh_type = "大型"
    veh_grade = "高一级"
    total_weight = "7900"
    fuel_type = "LNG"


    wb = openpyxl.load_workbook('P020210219545211473702.xlsx')
    ws = wb.active
    maxRows = ws.max_row
    maxColumn = ws.max_column

    # 定义对象
    VehicleType = collections.namedtuple("VehcleType",
                     ["product_no", "product_name", "mark", "oem", "veh_type", "veh_grade",
                      "total_weight", "fuel_type"])

    for row in range(3, maxRows + 1):
        product_no = ""
        product_name = ""
        mark = ""
        oem = ""
        veh_type = ""
        veh_grade = ""
        total_weight = ""
        fuel_type = ""
        requestUrl = ""
        for column in range(1, maxColumn + 1):
            if column > 5 and column < 27:
                continue
            columnValue = ws.cell(row, column).value
            print(str(column) + ": " + str(columnValue))
            if column == 2:
                product_no = columnValue
            elif column == 3:
                product_name = columnValue
            elif column == 5:
                oem = columnValue
            elif column == 29:
                if columnValue:
                    requestUrl += columnValue
                else:
                    continue
            elif column == 30:
                if columnValue:
                    requestUrl += columnValue
                else:
                    continue
        if requestUrl:
            requestUrl += ".html"
            print(requestUrl)
            response = requests.get(requestUrl)
            if response.status_code == 200:
                page = response.text
                tree = html.fromstring(page)
                table = tree.xpath("//table")[1]
                veh_type_ele = table.xpath("//td")[17]
                veh_type = veh_type_ele.text
                veh_grade_ele = table.xpath("//td")[19]
                veh_grade = veh_grade_ele.text
                print(veh_grade)
                total_weight_ele = table.xpath("//td")[21]
                total_weight = total_weight_ele.text
                print(total_weight)
                fuel_type_ele = table.xpath("//td")[25]
                fuel_type = fuel_type_ele.text
                vehicleType = VehicleType(product_no, product_name, mark, oem, veh_type, veh_grade, total_weight, fuel_type)
                print(vehicleType)
                update_veh_type(vehicleType)
