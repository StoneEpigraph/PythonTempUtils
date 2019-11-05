#! /usr/bin/env python3
# coding: utf-8


from openpyxl import load_workbook
from  decimal import Decimal

if __name__ == '__main__':
    doc = load_workbook("途牛20946.84元附件明细.xlsx")
    # print(doc)
    sheets = doc.sheetnames
    # print(sheets)
    # print(dir(doc))

    one_sheet = doc[sheets[0]]
    two_sheet = doc[sheets[1]]
    three_sheet = doc[sheets[2]]
    # print(one_sheet)
    # print(two_sheet)
    # print(three_sheet)

    parent_value = dict()
    for index, row in enumerate(one_sheet):
        if index < 2:
            continue
        # print(row)
        # print(dir(row))
        key = row[3].value
        value = row[13].value
        # print(row[3].value)
        # print(row[13].value)
        # print(dir(row[3]))
        old_value = parent_value.get(key)
        # print(old_value)
        if not old_value:
            parent_value[key] = value
        else:
            parent_value[key] = Decimal(old_value) + Decimal(value)

    print(parent_value)

    for index, row in enumerate(two_sheet):
        if index < 2:
            continue
        key = row[3].value
        value = row[13].value
        if parent_value.get(key):
            old_value = parent_value[key]
            new_value = Decimal(old_value) - Decimal(value)
            parent_value[key] = new_value

    for index, row in enumerate(three_sheet):
        if index < 2:
            continue
        key = row[3].value
        value = row[13].value
        if parent_value.get(key):
            old_value = parent_value[key]
            new_value = Decimal(old_value) - Decimal(value)
            parent_value[key] = new_value

    print(parent_value)